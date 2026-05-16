import pandas as pd
import openpyxl
import os
from datetime import date, datetime
from scripts.utils import get_working_days, ensure_project_context
from scripts.wbs_config_manager import WBSConfigManager

"""
EVM分析エンジン (EVMAnalyst)

【クラスの役割】
本クラスは、WBSデータから正確なEVM指標（PV, EV, AC）を計算し、
プロジェクトの現状診断と将来予測を行うための「正確な計算機」として動作します。

【設計思想：正確な計算と将来の拡張性】
1. 役割ベースのマッピング: 固定のカラム名ではなく、WBSConfigManager 経由で役割に応じた列を特定します。
2. 柔軟なフェーズ構造: テンプレートのフェーズ数に依存せず、存在する全てのフェーズを分析対象とします。
"""

class EVMAnalyst:
    def __init__(self, file_path, status_date=None, interactive=False):
        """
        EVM分析エンジンの初期化
        """
        self.file_path = file_path
        self.status_date = status_date if status_date else date.today()
        if isinstance(self.status_date, datetime):
            self.status_date = self.status_date.date()

        self.project_path = os.path.dirname(os.path.abspath(file_path))
        self.metadata = {
            "project_path": self.project_path,
            "context_path": os.path.join(self.project_path, "docs", "context.md")
        }
        
        self.config_manager = WBSConfigManager(file_path)
        self.config = self.config_manager.load_or_infer(interactive=interactive)

    def calculate_pv(self, start_date, end_date, planned_effort):
        """
        PV（Planned Value：計画値）を線形按分法で算出する。
        """
        if hasattr(start_date, 'date'): start_date = start_date.date()
        if hasattr(end_date, 'date'): end_date = end_date.date()
        
        if self.status_date < start_date:
            return 0.0
        if self.status_date >= end_date:
            return planned_effort
            
        total_working_days = get_working_days(start_date, end_date)
        if total_working_days <= 0:
            return 0.0
            
        elapsed_working_days = get_working_days(start_date, self.status_date)
        pv = (elapsed_working_days / total_working_days) * planned_effort
        return round(pv, 2)

    def calculate_ev(self, planned_effort, progress_rate):
        """
        EV（Earned Value：出来高）を算出する。
        """
        if pd.isna(progress_rate):
            return 0.0
        return round(planned_effort * (float(progress_rate) / 100.0), 2)

    def calculate_ac(self, actual_effort):
        """
        AC（Actual Cost：実績コスト）を算出する。
        """
        if pd.isna(actual_effort):
            return 0.0
        return float(actual_effort)

    def calculate_bac(self, df):
        """
        BAC（Budget At Completion：完成時総予算）を算出する。
        """
        total_bac = 0.0
        cm = self.config_manager
        for i in range(len(self.config["columns"]["phases"])):
            try:
                idx = cm.get_column_index("plan_effort", i)
                total_bac += df.iloc[:, idx].sum()
            except ValueError:
                continue
        return round(total_bac, 2)

    def calculate_forecasts(self, bac, ev, ac, pv):
        """
        PMBOK公式に基づき、3つの予測シナリオを算出する。
        """
        cpi = ev / ac if ac > 0 else 1.0
        spi = ev / pv if pv > 0 else 1.0
        
        safe_cpi = max(cpi, 0.01)
        safe_spi = max(spi, 0.01)
        
        eac_r = bac / safe_cpi
        eac_o = ac + (bac - ev)
        eac_p = ac + ((bac - ev) / (safe_cpi * safe_spi))
        
        def build_res(eac, desc):
            return {
                "eac": round(eac, 2),
                "etc": round(max(eac - ac, 0), 2),
                "vac": round(bac - eac, 2),
                "description": desc
            }
            
        return {
            "realistic": build_res(eac_r, "現在の効率が継続した場合の予測"),
            "optimistic": build_res(eac_o, "残作業を計画通りに遂行した場合"),
            "pessimistic": build_res(eac_p, "現在の効率と遅延が相互に影響した場合")
        }

    def run(self):
        """
        分析のメイン実行フロー。
        """
        from scripts.check_wbs_integrity import WBSIntegrityChecker
        
        print(f"--- 整合性チェックを開始します: {self.file_path} ---")
        checker = WBSIntegrityChecker(self.file_path, interactive=False) # 分析時は非対話
        df = checker.load_wbs()
        errors = checker.check_dataframe(df)
        
        if errors:
            checker.errors = errors
            checker.write_results_to_excel()
            raise ValueError(f"整合性チェックでエラーが検出されました ({len(errors)}件)。修正してから再度実行してください。")
        
        print("整合性チェックOK。分析を開始します...")

        results = []
        cm = self.config_manager
        phases = self.config["columns"]["phases"]
        idx_id = cm.get_column_index("id")
        start_row = self.config["data_start_row"]

        for index, row in df.iterrows():
            if pd.isna(row.iloc[idx_id]):
                continue

            for phase_idx in range(len(phases)):
                try:
                    s_idx = cm.get_column_index("plan_start", phase_idx)
                    e_idx = cm.get_column_index("plan_end", phase_idx)
                    m_idx = cm.get_column_index("plan_effort", phase_idx)
                    pr_idx = cm.get_column_index("progress", phase_idx)
                    act_idx = cm.get_column_index("actual_effort", phase_idx)
                    pv_idx = cm.get_column_index("pv", phase_idx)

                    if pd.isna(row.iloc[s_idx]) or pd.isna(row.iloc[e_idx]):
                        continue

                    pv = self.calculate_pv(row.iloc[s_idx], row.iloc[e_idx], row.iloc[m_idx])
                    ev = self.calculate_ev(row.iloc[m_idx], row.iloc[pr_idx])
                    ac = self.calculate_ac(row.iloc[act_idx])

                    results.append({
                        "row": index + start_row,
                        "pv": pv,
                        "ev": ev,
                        "ac": ac,
                        "excel_pv": row.iloc[pv_idx],
                        "phase_idx": phase_idx
                    })
                except ValueError:
                    continue

        summary = self.get_summary_json(results, df)
        import json
        print("\n--- 分析集計結果 (JSON) ---")
        print(json.dumps(summary, indent=2, ensure_ascii=False))
        print("---------------------------\n")
        
        print(f"分析が正常に完了しました。エクセルの数式は保護されました。")
        return summary

    def write_results_to_excel(self, results):
        """
        計算結果（PV, EV, AC）をExcelの該当セルに書き込む。
        """
        wb = openpyxl.load_workbook(self.file_path)
        ws = wb[self.config["sheet_name"]]
        cm = self.config_manager
        
        for res in results:
            row_idx = res["row"]
            p_idx = res["phase_idx"]
            
            try:
                ws.cell(row=row_idx, column=cm.get_column_index("pv", p_idx) + 1).value = res["pv"]
                ws.cell(row=row_idx, column=cm.get_column_index("ev", p_idx) + 1).value = res["ev"]
                ws.cell(row=row_idx, column=cm.get_column_index("ac", p_idx) + 1).value = res["ac"]
            except ValueError:
                continue

        wb.save(self.file_path)
        print(f"Excelへの計算結果反映が完了しました: {self.file_path}")

    def get_summary_json(self, results, df=None):
        """
        全体の集計を行い、AIアナリスト向けのJSONデータを生成する。
        """
        total_pv = sum(r["pv"] for r in results)
        total_ev = sum(r["ev"] for r in results)
        total_ac = sum(r["ac"] for r in results)
        
        gaps = []
        for r in results:
            try:
                e_pv = float(r["excel_pv"]) if not pd.isna(r["excel_pv"]) else 0.0
                diff = abs(r["pv"] - e_pv)
                if diff > 0.01:
                    gaps.append({"row": r["row"], "excel_pv": e_pv, "python_pv": r["pv"], "diff": diff})
            except (ValueError, TypeError):
                continue

        cpi = total_ev / total_ac if total_ac > 0 else 1.0
        spi = total_ev / total_pv if total_pv > 0 else 1.0
        
        summary = {
            "status_date": str(self.status_date),
            "metadata": self.metadata,
            "metrics": {
                "total_pv": round(total_pv, 2),
                "total_ev": round(total_ev, 2),
                "total_ac": round(total_ac, 2),
                "cpi": round(cpi, 2),
                "spi": round(spi, 2)
            },
            "gap_analysis": {
                "count": len(gaps),
                "details": gaps[:5]
            },
            "alerts": []
        }
        
        if df is not None:
            bac = self.calculate_bac(df)
            forecasts = self.calculate_forecasts(bac, total_ev, total_ac, total_pv)
            summary["forecasting"] = {
                "bac": bac,
                "scenarios": forecasts
            }

        if cpi < 0.9: summary["alerts"].append("COST_EFFICIENCY_LOW")
        if spi < 0.9: summary["alerts"].append("SCHEDULE_DELAYED")
        if len(gaps) > 0: summary["alerts"].append("EXCEL_FORMULA_GAP_DETECTED")
        
        return summary

def analyze_project(file_path, status_date=None, interactive=False):
    """
    プロジェクトの分析を実行するエントリーポイント関数。
    """
    project_path = os.path.dirname(os.path.abspath(file_path))
    ensure_project_context(project_path)
    
    analyst = EVMAnalyst(file_path, status_date=status_date, interactive=interactive)
    return analyst.run()

if __name__ == "__main__":
    import sys
    import argparse
    from datetime import datetime

    parser = argparse.ArgumentParser(description='WBS/EVM 自動分析エンジン')
    parser.add_argument('file_path', help='分析対象のエクセルファイルパス')
    parser.add_argument('--date', help='分析基準日 (YYYY-MM-DD)。指定がない場合は今日。')
    parser.add_argument('--non-interactive', action='store_true', help='対話型プロンプトを無効化する')
    
    args = parser.parse_args()
    
    status_date_val = None
    if args.date:
        try:
            status_date_val = datetime.strptime(args.date, '%Y-%m-%d').date()
        except ValueError:
            print("エラー: 日付形式は YYYY-MM-DD で指定してください。")
            sys.exit(1)
    
    analyze_project(args.file_path, status_date=status_date_val, interactive=not args.non_interactive)
