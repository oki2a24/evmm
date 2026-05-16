import pytest
import os
import pandas as pd
import openpyxl
import shutil
from datetime import date, datetime
from scripts.analyze_evm import EVMAnalyst, analyze_project

@pytest.fixture
def dummy_excel(tmp_path):
    """推論ロジックが通る程度の最低限のヘッダーを持つエクセルを作成"""
    file_path = tmp_path / "dummy.xlsx"
    df = pd.DataFrame([
        ["作成", None, None, None, None, None, None, None, None, None],
        ["機能ID", "機能名称", "開始日予定", "終了日予定", "工数予定", "開始日実績", "終了日実績", "工数実績", "進捗率(%)", "PV (計画値)"],
        ["F1", "Test", "2026-05-01", "2026-05-10", 10.0, None, None, 0.0, 0.0, 0.0]
    ])
    df.to_excel(file_path, index=False, header=False, sheet_name="WBS_EVM")
    return str(file_path)

def test_analyze_project_integrates_context_initialization(mocker, tmp_path, dummy_excel):
    """
    analyze_project を実行した際、プロジェクトコンテキストの初期化が行われ、
    メタデータにパスが含まれることを検証する。
    """
    # ensure_project_context をモック
    mock_ensure = mocker.patch("scripts.analyze_evm.ensure_project_context")
    
    # ダミーの WBS データを返すように EVMAnalyst.run をモック
    mocker.patch.object(EVMAnalyst, "run", return_value={
        "status_date": "2026-05-10",
        "metadata": {"context_path": str(os.path.dirname(dummy_excel) + "/docs/context.md")},
        "metrics": {"total_pv": 10.0}
    })
    
    result = analyze_project(dummy_excel)
    
    # ensure_project_context が正しいプロジェクトパスで呼ばれたか
    mock_ensure.assert_called_once_with(os.path.dirname(dummy_excel))
    
    # メタデータに context_path が含まれているか
    assert "metadata" in result
    assert result["metrics"]["total_pv"] == 10.0

def test_calculate_pv_linear(dummy_excel):
    """
    PVの線形按分計算のテスト（正常系：期間中）
    """
    analyst = EVMAnalyst(file_path=dummy_excel, status_date=date(2026, 4, 7))
    
    start_date = date(2026, 4, 1)
    end_date = date(2026, 4, 14)
    planned_effort = 10.0
    
    pv = analyst.calculate_pv(start_date, end_date, planned_effort)
    assert pv == 5.0

def test_calculate_pv_before_start(dummy_excel):
    """基準日が開始日より前の場合、PVは 0.0 であるべき"""
    analyst = EVMAnalyst(file_path=dummy_excel, status_date=date(2026, 3, 31))
    pv = analyst.calculate_pv(date(2026, 4, 1), date(2026, 4, 10), 10.0)
    assert pv == 0.0

def test_calculate_pv_after_end(dummy_excel):
    """基準日が終了日より後の場合、PVは全額（工数予定）であるべき"""
    analyst = EVMAnalyst(file_path=dummy_excel, status_date=date(2026, 4, 15))
    pv = analyst.calculate_pv(date(2026, 4, 1), date(2026, 4, 10), 10.0)
    assert pv == 10.0

def test_calculate_ev(dummy_excel):
    """EV = 工数予定 * 進捗率"""
    analyst = EVMAnalyst(file_path=dummy_excel)
    ev = analyst.calculate_ev(10.0, 50.0)
    assert ev == 5.0
    
    ev_zero = analyst.calculate_ev(10.0, 0.0)
    assert ev_zero == 0.0

def test_calculate_ac(dummy_excel):
    """AC = 工数実績の値そのもの"""
    analyst = EVMAnalyst(file_path=dummy_excel)
    ac = analyst.calculate_ac(8.5)
    assert ac == 8.5
    
    ac_none = analyst.calculate_ac(None)
    assert ac_none == 0.0

def test_run_with_integrity_error(mocker, dummy_excel):
    """
    整合性エラーがある場合に分析を中断することを検証。
    """
    from scripts.check_wbs_integrity import WBSIntegrityChecker
    
    # WBSIntegrityChecker.check_dataframe がエラーを返すようにモック
    mocker.patch.object(WBSIntegrityChecker, 'load_wbs', return_value=None)
    mocker.patch.object(WBSIntegrityChecker, 'check_dataframe', return_value=[{"index": 0, "message": "Error"}])
    mocker.patch.object(WBSIntegrityChecker, 'write_results_to_excel', return_value=None)
    
    analyst = EVMAnalyst(file_path=dummy_excel)
    
    # 整合性エラーがある場合、ValueError を送出することを期待
    with pytest.raises(ValueError, match="整合性チェックでエラーが検出されました"):
        analyst.run()

def test_write_results_to_excel(tmp_path):
    """
    Excelへの書き出しが正しく行われるかを検証。
    """
    template_path = "tests/data/wbs_template_for_testing.xlsx"
    test_excel = tmp_path / "test_analyze.xlsx"
    shutil.copy(template_path, test_excel)
    
    analyst = EVMAnalyst(file_path=str(test_excel), status_date=date(2026, 4, 7))
    
    # phase_idx を追加
    results = [
        {"row": 3, "pv": 5.0, "ev": 2.0, "ac": 1.5, "phase_idx": 0}
    ]
    
    analyst.write_results_to_excel(results)
    
    wb = openpyxl.load_workbook(str(test_excel))
    ws = wb['WBS_EVM']
    
    # テンプレート構造: PV=13列, EV=14列, AC=15列
    assert ws.cell(row=3, column=13).value == 5.0
    assert ws.cell(row=3, column=14).value == 2.0
    assert ws.cell(row=3, column=15).value == 1.5

def test_calculate_bac(tmp_path):
    """
    BAC (完成時総予算) 算出のテスト。
    """
    template_path = "tests/data/wbs_template_for_testing.xlsx"
    test_excel = tmp_path / "test_bac.xlsx"
    shutil.copy(template_path, test_excel)

    wb = openpyxl.load_workbook(str(test_excel))
    ws = wb['WBS_EVM']
    # 3行目: 作成(8列目)=10, レビュー(20列目)=2, 修正(32列目)=1 -> 計13
    ws.cell(row=3, column=8, value=10.0)
    ws.cell(row=3, column=20, value=2.0)
    ws.cell(row=3, column=32, value=1.0)
    # 4行目: 作成(8列目)=5 -> 計18
    ws.cell(row=4, column=8, value=5.0)
    wb.save(str(test_excel))

    analyst = EVMAnalyst(file_path=str(test_excel))
    df = analyst.config_manager.load_or_infer(interactive=False) # ダミーでロード
    # 実際には analyst.run 等で使われる DataFrame を渡す必要がある
    from scripts.check_wbs_integrity import WBSIntegrityChecker
    checker = WBSIntegrityChecker(str(test_excel), interactive=False)
    df = checker.load_wbs()
    bac = analyst.calculate_bac(df)
    
    assert bac == 18.0

def test_calculate_forecasts(dummy_excel):
    """
    将来予測（3シナリオ）の計算テスト。
    """
    analyst = EVMAnalyst(file_path=dummy_excel)
    
    bac = 100.0
    ev = 40.0
    ac = 50.0
    pv = 50.0
    
    forecasts = analyst.calculate_forecasts(bac, ev, ac, pv)
    
    res = forecasts["realistic"]
    assert res["eac"] == 125.0
    opt = forecasts["optimistic"]
    assert opt["eac"] == 110.0
    pes = forecasts["pessimistic"]
    assert pes["eac"] == 143.75

def test_calculate_forecasts_zero_efficiency(dummy_excel):
    """効率が0の場合のガードレール検証"""
    analyst = EVMAnalyst(file_path=dummy_excel)
    forecasts = analyst.calculate_forecasts(bac=100.0, ev=0.0, ac=10.0, pv=10.0)
    assert forecasts["realistic"]["eac"] >= 1000.0
    assert forecasts["optimistic"]["eac"] == 110.0
