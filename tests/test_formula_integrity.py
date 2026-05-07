import pytest
from scripts.utils import FormulaIntegrityManager

def test_extract_formula_template_relative():
    """相対参照を含む数式からテンプレートを抽出する"""
    manager = FormulaIntegrityManager()
    # 行3の数式を抽象化
    formula = "=B3+C3"
    template = manager.extract_template(formula, row=3)
    assert template == "={B}{row}+{C}{row}"

def test_extract_formula_template_absolute():
    """絶対参照を含む数式は抽象化しない"""
    manager = FormulaIntegrityManager()
    formula = "=$A$1+B3"
    template = manager.extract_template(formula, row=3)
    assert template == "=$A$1+{B}{row}"

def test_extract_formula_template_complex():
    """関数や複雑な参照を含む数式"""
    manager = FormulaIntegrityManager()
    formula = "=IF(TODAY()<D3, 0, E3*100)"
    template = manager.extract_template(formula, row=3)
    assert template == "=IF(TODAY()<{D}{row}, 0, {E}{row}*100)"

def test_extract_formula_lowercase():
    """小文字の数式も大文字に正規化してテンプレート化する"""
    manager = FormulaIntegrityManager()
    formula = "=b3+c3"
    template = manager.extract_template(formula, row=3)
    assert template == "={B}{row}+{C}{row}"

def test_extract_formula_sheet_ref():
    """他シート参照（!付き）は抽象化しない（自シート内の行一貫性のみを対象とするため）"""
    manager = FormulaIntegrityManager()
    formula = "=Sheet1!A3+B3"
    template = manager.extract_template(formula, row=3)
    assert template == "=Sheet1!A3+{B}{row}"

def test_extract_formula_multi_char_col():
    """AA列などの複数文字の列名に対応"""
    manager = FormulaIntegrityManager()
    formula = "=AA3+AB3"
    template = manager.extract_template(formula, row=3)
    assert template == "={AA}{row}+{AB}{row}"

def test_extract_formula_no_false_positive():
    """VALUE3 などの単語の末尾をセル参照(LUE3)と誤認しないこと"""
    manager = FormulaIntegrityManager()
    # VALUE3 の末尾 LUE3 がマッチしてしまうのを防ぎたい
    formula = '=IF(A3="VALUE3", B3, 0)'
    template = manager.extract_template(formula, row=3)
    assert template == '=IF({A}{row}="VALUE3", {B}{row}, 0)'

from openpyxl import Workbook

def test_repair_sheet_majority_vote():
    """多数決により正解パターンが選ばれ、破損（数値上書き）が修復されることを検証する"""
    wb = Workbook()
    ws = wb.active
    ws.title = "WBS_EVM"
    
    # 3行目: 正解
    ws["A3"] = "=B3+C3"
    # 4行目: 正解
    ws["A4"] = "=B4+C4"
    # 5行目: 破損（数値上書き）
    ws["A5"] = 100 
    # 6行目: 異ロジック（少数派）
    ws["A6"] = "=B6*D6"
    
    manager = FormulaIntegrityManager()
    # A列（列番号1）をスキャンして修復
    # データ行は 3〜6行目
    repaired_count = manager.repair_sheet(ws, columns=[1], start_row=3, end_row=6)
    
    assert repaired_count == 2 # 5行目と6行目が修復されるはず
    assert ws["A5"].value == "=B5+C5"
    assert ws["A6"].value == "=B6+C6"

def test_repair_sheet_no_majority():
    """多数決でタイの場合、安全のため修復しない"""
    wb = Workbook()
    ws = wb.active
    ws["A3"] = "=B3+C3"
    ws["A4"] = "=B4*D4"
    
    manager = FormulaIntegrityManager()
    repaired_count = manager.repair_sheet(ws, columns=[1], start_row=3, end_row=4)
    
    assert repaired_count == 0
    assert ws["A3"].value == "=B3+C3"
    assert ws["A4"].value == "=B4*D4"
