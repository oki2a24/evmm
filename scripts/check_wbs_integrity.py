import pandas as pd
import openpyxl
import re
from collections import Counter
from scripts.utils import get_working_days, FormulaIntegrityManager
from scripts.wbs_config_manager import WBSConfigManager

class WBSIntegrityChecker:
    """
    WBSの論理的整合性をチェックし、必要に応じて数式を修復するクラス。
    """
    def __init__(self, file_path, fix=False, interactive=False):
        self.file_path = file_path
        self.fix = fix
        self.interactive = interactive
        self.errors = []
        self.formula_manager = FormulaIntegrityManager()
        self.config_manager = WBSConfigManager(file_path)
        self.config = self.config_manager.load_or_infer(interactive=interactive)

    def load_wbs(self):
        """
        ExcelからWBSデータを読み込む。
        """
        df = pd.read_excel(
            self.file_path, 
            sheet_name=self.config["sheet_name"], 
            skiprows=self.config["header_row"] - 1
        )
        return df

    def check_dataframe(self, df):
        """
        DataFrameに対して整合性チェックを実行する。
        """
        errors = []
        cm = self.config_manager
        phases = self.config["columns"]["phases"]
        
        idx_id = cm.get_column_index("id")

        for index, row in df.iterrows():
            if idx_id is not None and pd.isna(row.iloc[idx_id]):
                continue

            # --- 1. 日付の前後関係と工数超過チェック ---
            # 【堅牢な設計 (Robust Design)】
            # 特定の項目（例：工数予定）がマッピングに存在しなくても、可能な範囲で
            # 日付のチェック等を続行します。これにより、全ての項目が埋まっていない
            # 開発途中の WBS に対しても有益なフィードバックを提供できます。
            prev_end = None
            for i, p in enumerate(phases):
                # 個別の項目を安全に取得
                def get_idx(role):
                    try: return cm.get_column_index(role, i)
                    except ValueError: return None


                s_idx = get_idx("plan_start")
                e_idx = get_idx("plan_end")
                m_idx = get_idx("plan_effort")
                
                s_p = pd.to_datetime(row.iloc[s_idx]) if s_idx is not None and not pd.isna(row.iloc[s_idx]) else pd.NaT
                e_p = pd.to_datetime(row.iloc[e_idx]) if e_idx is not None and not pd.isna(row.iloc[e_idx]) else pd.NaT
                m_p = row.iloc[m_idx] if m_idx is not None and not pd.isna(row.iloc[m_idx]) else None

                if s_idx is not None and pd.isna(s_p):
                    errors.append({"index": index, "message": f"{p['name']}フェーズの開始日予定が未入力です。"})
                if e_idx is not None and pd.isna(e_p):
                    errors.append({"index": index, "message": f"{p['name']}フェーズの終了日予定が未入力です。"})

                if not pd.isna(s_p) and not pd.isna(e_p):
                    if s_p > e_p:
                        errors.append({"index": index, "message": f"{p['name']}フェーズの開始日予定が終了日予定より後になっています({s_p.date()} > {e_p.date()})。"})
                    
                    if m_p is not None:
                        w_days = get_working_days(s_p.date(), e_p.date())
                        if m_p > w_days:
                            errors.append({"index": index, "message": f"{p['name']}フェーズの工数予定({m_p})が稼働日数({w_days})を超過しています(過負荷)。"})

                # 前フェーズとの順序関係 (Finish-to-Start)
                if prev_end and not pd.isna(s_p):
                    if prev_end > s_p:
                        errors.append({"index": index, "message": f"前フェーズの終了日({prev_end.date()})より前に{p['name']}フェーズが開始({s_p.date()})されています。"})
                
                if not pd.isna(e_p):
                    prev_end = e_p



            # --- 2. 進捗率と終了実績の整合性 ---
            try:
                pr_idx = cm.get_column_index("progress", 0)
                ea_idx = cm.get_column_index("actual_end", 0)
                progress = row.iloc[pr_idx]
                e_a = pd.to_datetime(row.iloc[ea_idx]) if not pd.isna(row.iloc[ea_idx]) else pd.NaT
                
                if progress == 100.0 and pd.isna(e_a):
                    errors.append({"index": index, "message": "進捗率100%ですが、終了日実績が未入力です。"})
            except (ValueError, IndexError):
                pass

        # --- 3. 数式の整合性チェック ---
        formula_cols = []
        for i, p in enumerate(phases):
            for role in ["pv", "ev", "ac"]:
                try:
                    formula_cols.append(cm.get_column_index(role, i))
                except ValueError:
                    continue

        if formula_cols:
            wb = openpyxl.load_workbook(self.file_path, data_only=False)
            ws = wb[self.config["sheet_name"]]
            start_row = self.config["data_start_row"]
            
            for col_idx in formula_cols:
                patterns = []
                for row in range(start_row, ws.max_row + 1):
                    cell_val = ws.cell(row=row, column=col_idx+1).value
                    template = self.formula_manager.extract_template(cell_val, row)
                    if template:
                        patterns.append(template)
                
                if patterns:
                    counter = Counter(patterns)
                    most_common = counter.most_common(2)
                    winner_pattern, count = most_common[0]
                    
                    if not (len(most_common) > 1 and most_common[0][1] == most_common[1][1]):
                        for row in range(start_row, ws.max_row + 1):
                            cell_val = ws.cell(row=row, column=col_idx+1).value
                            current_template = self.formula_manager.extract_template(cell_val, row)
                            if current_template != winner_pattern:
                                col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                                errors.append({
                                    "index": row - start_row, 
                                    "message": f"数式の不整合を検知しました (列 {col_letter})。期待値テンプレート: {winner_pattern}"
                                })
            wb.close()

        return errors

    def write_results_to_excel(self):
        """
        チェック結果をExcelの「整合性チェック結果」列に書き戻す。
        """
        wb = openpyxl.load_workbook(self.file_path)
        ws = wb[self.config["sheet_name"]]
        header_row = self.config["header_row"]
        start_row = self.config["data_start_row"]
        last_col = ws.max_column
        target_col = None

        for c in range(1, last_col + 1):
            if ws.cell(row=header_row, column=c).value == "整合性チェック結果":
                target_col = c
                break

        if target_col is None:
            target_col = last_col + 1
            ws.cell(row=header_row, column=target_col).value = "整合性チェック結果"
            target_cell = ws.cell(row=header_row, column=target_col)
            target_cell.font = openpyxl.styles.Font(bold=True)
            target_cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            target_cell.fill = openpyxl.styles.PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        for r in range(start_row, ws.max_row + 1):
            ws.cell(row=r, column=target_col).value = None

        for e in self.errors:
            excel_row = e['index'] + start_row
            current_val = ws.cell(row=excel_row, column=target_col).value
            new_msg = e['message']
            if current_val:
                new_msg = f"{current_val} | {new_msg}"
            ws.cell(row=excel_row, column=target_col).value = new_msg
            ws.cell(row=excel_row, column=target_col).font = openpyxl.styles.Font(color="FF0000")

        wb.save(self.file_path)
        print(f"Excelにチェック結果を書き込みました: {self.file_path}")

    def run(self):
        """
        実行メイン処理
        """
        df = self.load_wbs()
        self.errors = self.check_dataframe(df)

        if not self.errors:
            print("整合性チェック完了: 問題は見つかりませんでした。")
        else:
            print(f"整合性チェック完了: {len(self.errors)} 件の問題が見つかりました。")
            start_row = self.config["data_start_row"]
            for e in self.errors:
                print(f"行 {e['index'] + start_row}: {e['message']}")
            
            if not self.fix:
                print("\n[HINT] 数式の不整合が検知されました。'scripts/check_wbs.sh --fix' を実行することで、多数決パターンに基づき数式を自動修復できます。")

        if self.fix and self.errors:
            import shutil
            bak_path = self.file_path + ".bak"
            shutil.copy2(self.file_path, bak_path)
            print(f"バックアップを作成しました: {bak_path}")

            wb = openpyxl.load_workbook(self.file_path)
            ws = wb[self.config["sheet_name"]]
            start_row = self.config["data_start_row"]
            
            formula_cols = []
            for i, p in enumerate(self.config["columns"]["phases"]):
                for role in ["pv", "ev", "ac"]:
                    try:
                        formula_cols.append(self.config_manager.get_column_index(role, i))
                    except ValueError:
                        continue
            
            repaired = self.formula_manager.repair_sheet(
                ws, 
                columns=[c+1 for c in formula_cols], 
                start_row=start_row, 
                end_row=ws.max_row
            )
            
            if repaired > 0:
                wb.save(self.file_path)
                print(f"成功: {repaired} 個のセルを修復しました。")
                print("修復後の再検証を実行中...")
                new_df = self.load_wbs()
                self.errors = self.check_dataframe(new_df)
                if not self.errors:
                    print("再検証の結果、全ての不整合が解消されました。")
            else:
                print("修復可能な不整合は見つかりませんでした。")

        self.write_results_to_excel()

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description='WBS整合性チェック・修復スクリプト')
    parser.add_argument('file', help='Excelファイルパス')
    parser.add_argument('--fix', action='store_true', help='数式の不整合を自動修復する')
    parser.add_argument('--non-interactive', action='store_true', help='対話型プロンプトを無効化する')
    
    args = parser.parse_args()
    checker = WBSIntegrityChecker(args.file, fix=args.fix, interactive=not args.non_interactive)
    checker.run()
