import pytest
import openpyxl
import pandas as pd
import shutil
import os
from datetime import datetime
from scripts.update_wbs import update_wbs_logic

# テスト用のダミーWBSパス
TEST_EXCEL = "tests/test_data_update.xlsx"
# 本番データ(hoge_wbs_evm.xlsx)に依存せず、固定のテスト用テンプレートを使用する
TEMPLATE_EXCEL = "tests/data/wbs_template_for_testing.xlsx"

@pytest.fixture
def setup_excel():
    """テスト実行前にテンプレートをコピーして環境を整える"""
    if not os.path.exists("tests"):
        os.makedirs("tests")
    shutil.copy(TEMPLATE_EXCEL, TEST_EXCEL)
    yield TEST_EXCEL
    # テスト後に削除
    if os.path.exists(TEST_EXCEL):
        os.remove(TEST_EXCEL)

def test_update_wbs_basic(setup_excel):
    """
    基本機能のテスト: F1の『作成』フェーズを工数3hで更新する。
    """
    # 実行前の値を確認
    wb = openpyxl.load_workbook(setup_excel)
    ws = wb['WBS_EVM']
    # F1は3行目
    original_effort = ws.cell(row=3, column=11).value
    wb.close()

    # 更新実行
    new_effort = 3.0
    update_wbs_logic(setup_excel, func_id="F1", phase="作成", effort=new_effort)

    # 実行後の値を確認
    wb = openpyxl.load_workbook(setup_excel)
    ws = wb['WBS_EVM']
    updated_effort = ws.cell(row=3, column=11).value
    updated_progress = ws.cell(row=3, column=12).value
    wb.close()

    assert updated_effort == new_effort
    assert updated_progress == 100

def test_update_wbs_stateful(setup_excel):
    """
    ステートフル更新のテスト: 
    1. F1の『作成』を100%にする。
    2. phaseを指定せずに呼び出した場合、自動的に次の『レビュー実施』が更新されることを確認する。
    """
    # 1. まず『作成』を完了させる
    update_wbs_logic(setup_excel, func_id="F1", phase="作成", effort=2.0)
    
    # 2. 次に phase を None で呼び出す
    new_effort_review = 1.5
    update_wbs_logic(setup_excel, func_id="F1", phase=None, effort=new_effort_review)

    # 3. 『レビュー実施』の列(occurrence=1)が更新されているか確認
    # レビュー実施フェーズの工数実績は W列(23列目)
    wb = openpyxl.load_workbook(setup_excel)
    ws = wb['WBS_EVM']
    updated_effort_review = ws.cell(row=3, column=23).value
    wb.close()

    assert updated_effort_review == new_effort_review

def test_update_wbs_atomicity(setup_excel):
    """
    原子性のテスト: 
    整合性エラーが発生するような状態で更新を試みた際、
    例外が投げられ、かつ元のファイルが書き換わっていないことを確認する。
    """
    # 1. 事前にファイルを『壊れた』状態（開始日予定 > 終了日予定）にする
    wb = openpyxl.load_workbook(setup_excel)
    ws = wb['WBS_EVM']
    ws.cell(row=3, column=6, value=datetime(2026, 12, 31))
    ws.cell(row=3, column=7, value=datetime(2026, 1, 1))
    
    original_effort = ws.cell(row=3, column=11).value
    wb.save(setup_excel)
    wb.close()

    # 2. 更新を試みる。
    with pytest.raises(ValueError) as excinfo:
        update_wbs_logic(setup_excel, func_id="F1", phase="作成", effort=99.0)
    
    assert "整合性チェックエラー" in str(excinfo.value)

    # 3. ファイルの工数実績が書き換わっていないことを確認
    wb = openpyxl.load_workbook(setup_excel)
    ws = wb['WBS_EVM']
    final_effort = ws.cell(row=3, column=11).value
    wb.close()

    assert final_effort == original_effort

def test_update_wbs_with_date_and_progress(setup_excel):
    """
    新機能のテスト: 過去の日付指定および進捗率の指定。
    F4のレビュー実施フェーズを 4/3 に 4.0h、進捗50%で更新する。
    """
    past_date = datetime(2026, 4, 3)
    target_progress = 50.0
    target_effort = 4.0

    update_wbs_logic(setup_excel, func_id="F4", phase="レビュー実施", effort=target_effort, date=past_date, progress=target_progress)

    wb = openpyxl.load_workbook(setup_excel)
    ws = wb['WBS_EVM']
    # V=22, W=23, X=24
    cell_val = ws.cell(row=6, column=22).value
    # openpyxl がシリアル値(int)で返す場合は変換する
    if isinstance(cell_val, int):
        from openpyxl.utils.datetime import from_excel
        updated_date = from_excel(cell_val)
    else:
        updated_date = cell_val

    updated_effort = ws.cell(row=6, column=23).value
    updated_progress = ws.cell(row=6, column=24).value
    wb.close()

    assert updated_date.date() == past_date.date()
    assert updated_effort == target_effort
    assert updated_progress == target_progress

def test_update_wbs_date_format(setup_excel):
    """
    書式設定のテスト:
    日付を更新した際、セルの表示形式 (number_format) が継承されていることを確認する。
    """
    # 更新実行
    update_wbs_logic(setup_excel, func_id="F1", phase="レビュー実施", effort=1.0)

    wb = openpyxl.load_workbook(setup_excel)
    ws = wb['WBS_EVM']
    
    # レビュー実施フェーズ：開始日実績(21列), 終了日実績(22列)
    cell_sa = ws.cell(row=3, column=21) 
    cell_ea = ws.cell(row=3, column=22) 
    
    # テンプレートで設定した期待されるフォーマット (m\月d\日)
    expected_format = 'm\\月d\\日'
    
    fmt_sa = cell_sa.number_format
    fmt_ea = cell_ea.number_format
    wb.close()

    assert fmt_sa == expected_format, f"開始日実績の書式が不正です: {fmt_sa}"
    assert fmt_ea == expected_format, f"終了日実績の書式が不正です: {fmt_ea}"

def test_update_wbs_cli_help():
    """
    CLI インターフェースのテスト
    """
    import subprocess
    result = subprocess.run(
        [".venv/bin/python3", "scripts/update_wbs.py", "--help"],
        capture_output=True,
        text=True
    )
    assert result.returncode == 0
    assert "進捗率" in result.stdout
