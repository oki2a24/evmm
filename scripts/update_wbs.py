import pandas as pd
import openpyxl
import os
import tempfile
import shutil
from datetime import datetime
from scripts.check_wbs_integrity import WBSIntegrityChecker
from scripts.wbs_config_manager import WBSConfigManager

def update_wbs_logic(file_path, func_id, phase, effort, date=None, progress=100, interactive=True):
    """
    Excel WBSの特定の機能・フェーズを更新する。

    【設計の背景と意思決定 (Architecture & Decisions)】
    1. 役割ベースのマッピングへの移行:
       - 従来は「開始日予定」などの文字列を直接探していましたが、WBSConfigManager を導入し、
         役割（Role）に基づいて動的に列を特定する方式に変更しました。
    2. フェーズの動的解決:
       - フェーズ名も固定ではなく、エクセルから読み取った実際のフェーズリストに基づきます。
    3. 原子性の担保 (Atomic Integrity):
       - 一時ファイルに対して更新・検証（整合性チェック）を行い、成功した場合のみ元のファイルを上書き。
    """
    if date is None:
        date = datetime.now()

    # WBS構造の読み込み
    config_manager = WBSConfigManager(file_path)
    config = config_manager.load_or_infer(interactive=interactive)
    
    sheet_name = config["sheet_name"]
    header_row_idx = config["header_row"]
    data_start_row = config["data_start_row"]

    # 1. カラム位置の特定のために一度 pandas で読み込む
    df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=header_row_idx-1)

    # フェーズの自動特定
    phases = config["columns"]["phases"]
    if phase is None:
        idx_id_col = config_manager.get_column_index("id")
        
        target_row = df[df.iloc[:, idx_id_col] == func_id]
        if target_row.empty:
            raise ValueError(f"機能ID '{func_id}' が見つかりません。")
        
        row_data = target_row.iloc[0]
        
        # 未完了の最初のフェーズを探す
        target_phase_idx = 0
        for i, p in enumerate(phases):
            try:
                pr_idx = config_manager.get_column_index("progress", i)
                pr_val = row_data.iloc[pr_idx] if not pd.isna(row_data.iloc[pr_idx]) else 0
                if pr_val < 100:
                    target_phase_idx = i
                    break
            except ValueError:
                continue # 進捗率列がないフェーズはスキップ
        else:
            target_phase_idx = len(phases) - 1
        
        phase_idx = target_phase_idx
    else:
        # 指定されたフェーズ名からインデックスを探す
        phase_idx = -1
        for i, p in enumerate(phases):
            if p["name"] == phase:
                phase_idx = i
                break
        if phase_idx == -1:
            raise ValueError(f"フェーズ '{phase}' が見つかりません。")

    idx_id = config_manager.get_column_index("id")
    idx_sp = config_manager.get_column_index("plan_start", phase_idx)
    idx_sa = config_manager.get_column_index("actual_start", phase_idx)
    idx_ea = config_manager.get_column_index("actual_end", phase_idx)
    idx_ma = config_manager.get_column_index("actual_effort", phase_idx)
    idx_pr = config_manager.get_column_index("progress", phase_idx)

    target_row_idx = df[df.iloc[:, idx_id] == func_id].index
    if target_row_idx.empty:
        raise ValueError(f"機能ID '{func_id}' が見つかりません。")
    
    # Excelの物理行番号を算出
    excel_row = target_row_idx[0] + data_start_row

    # --- トランザクション処理 ---
    fd, temp_path = tempfile.mkstemp(suffix=".xlsx", dir=os.path.dirname(file_path))
    os.close(fd)
    
    try:
        shutil.copy2(file_path, temp_path)

        wb = openpyxl.load_workbook(temp_path)
        ws = wb[sheet_name]

        # 書式（表示形式）を予定日セルから継承する
        date_format = "m/d" # フォールバック
        if idx_sp is not None:
            date_format = ws.cell(row=excel_row, column=idx_sp + 1).number_format

        # 開始日実績の更新
        cell_sa = ws.cell(row=excel_row, column=idx_sa + 1)
        if cell_sa.value is None:
            cell_sa.value = date
            cell_sa.number_format = date_format
        
        # 終了日実績の更新
        cell_ea = ws.cell(row=excel_row, column=idx_ea + 1)
        cell_ea.value = date
        cell_ea.number_format = date_format
        
        ws.cell(row=excel_row, column=idx_ma + 1, value=effort)
        ws.cell(row=excel_row, column=idx_pr + 1, value=progress)

        wb.save(temp_path)
        wb.close()

        # 整合性チェック (Hard Gate)
        checker = WBSIntegrityChecker(temp_path)
        df_for_check = checker.load_wbs()
        errors = checker.check_dataframe(df_for_check)
        if errors:
            error_msg = "\n".join([f"行 {e['index']+data_start_row}: {e['message']}" for e in errors])
            raise ValueError(f"整合性チェックエラーが発生しました。更新を中断します（ファイルは保護されました）:\n{error_msg}")

        # 検証成功！上書き
        shutil.move(temp_path, file_path)
        return True

    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description='WBS更新スクリプト')
    parser.add_argument('--id', required=True, help='機能ID')
    parser.add_argument('--phase', help='フェーズ (作成/レビュー実施/レビュー後修正)。省略時は進捗から自動判別。')
    parser.add_argument('--effort', type=float, required=True, help='実績工数')
    parser.add_argument('--date', help='実績終了日 (YYYY-MM-DD)。省略時は今日。')
    parser.add_argument('--progress', type=float, default=100, help='進捗率(%%)。デフォルトは100。')
    parser.add_argument('--file', default='projects/test_project/hoge_wbs_evm.xlsx', help='Excelファイルパス')
    parser.add_argument('--non-interactive', action='store_true', help='対話型プロンプトを無効化する')
    
    args = parser.parse_args()
    try:
        update_date = datetime.strptime(args.date, '%Y-%m-%d') if args.date else None
        update_wbs_logic(args.file, args.id, args.phase, args.effort, date=update_date, progress=args.progress, interactive=not args.non_interactive)
        print(f"SUCCESS: {args.id} を更新しました。")
    except Exception as e:
        print(f"ERROR: {e}")
        exit(1)
