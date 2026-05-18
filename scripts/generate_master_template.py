import os
import argparse
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter

"""
WBS/EVM マスターテンプレート生成スクリプト

【重要：このスクリプトの存在意義】
このスクリプトは、単なるエクセル作成ツールではなく、本プロジェクトにおける「正しい管理構造の定義書（SSOT: Single Source of Truth）」です。
バイナリであるエクセルファイルそのものをマスターとせず、本スクリプトをマスターとすることで以下の利点を得ます：

1. 整合性の保証: 数式、名前定義、シート構造のミスを、コードレビューとテスト（TDD）で防ぐことができます。
2. 自己修復性: 運用中にエクセルが破損したり、数式が上書きされたりしても、本スクリプトを実行すれば即座に「完璧な状態」に復元可能です。
3. 自動化エンジン: 将来的には、このスクリプトがパラメータを受け取り、プロジェクト規模や期間に応じた動的なテンプレートを生成する「知能」となります。

【注意】
- 本スクリプトで生成されたエクセルの「構造」や「数式」を、エクセル上で直接修正しないでください。
- 構造の変更が必要な場合は、まず本スクリプトを修正し、再生成してください。
"""

class TemplateGenerator:
    """テンプレート生成を管理するクラス"""
    
    def __init__(self, output_path):
        self.output_path = output_path
        self.wb = Workbook()
        self.config = None # 動的設定保持用
        # デフォルトのシートを削除して新しく作成
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]

    def _apply_borders(self, ws, min_row, max_row, min_col, max_col, outer_medium=True):
        """
        指定された範囲に罫線を適用する。
        内部は細線(thin)、外枠とフェーズ境界は太線(medium)を使用する。
        """
        thin = Side(style='thin')
        medium = Side(style='medium')
        
        # フェーズ境界の特定 (共通項目と各フェーズの末尾列)
        boundary_cols = []
        if self.config:
            # 共通項目の最大列
            if self.config["columns"]["common"]:
                boundary_cols.append(max(info["index"] + 1 for info in self.config["columns"]["common"].values()))
            # 各フェーズの最大列
            for phase in self.config["columns"]["phases"]:
                indices = [info["index"] for info in phase["mapping"].values()]
                if indices:
                    boundary_cols.append(max(indices) + 1)
        else:
            # デフォルト時の固定位置
            boundary_cols = [3, 15, 27]

        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                # デフォルトは全周囲細線
                border_args = {
                    'top': thin, 'bottom': thin, 'left': thin, 'right': thin
                }
                
                # 外枠の処理
                if outer_medium:
                    if cell.row == min_row: border_args['top'] = medium
                    if cell.row == max_row: border_args['bottom'] = medium
                    if cell.column == min_col: border_args['left'] = medium
                    if cell.column == max_col: border_args['right'] = medium
                
                # フェーズ境界の処理
                # ユーザーの利便性のため、フェーズの区切りに太線を入れる
                if ws.title in ["WBS_EVM", self.config.get("sheet_name") if self.config else None]:
                    if cell.column in boundary_cols:
                        border_args['right'] = medium
                
                cell.border = Border(**border_args)

    def _create_settings_sheet(self):
        """Settingsシートを作成し、基本情報と名前定義を設定する"""
        ws = self.wb.create_sheet("Settings")
        
        header_font = Font(bold=True, size=12)
        
        ws["A1"] = "プロジェクト基本情報"
        ws["A1"].font = header_font
        
        ws["A2"] = "プロジェクト開始日"
        ws["B2"] = "2026/05/01"
        
        ws["A3"] = "メンバーリスト"
        ws["A3"].font = header_font
        
        ws["A4"] = "メンバー名"
        ws["B4"] = "役割"
        ws["C4"] = "チームリーダー"
        
        members = [
            ("Aさん", "リーダー", "Aさん"),
            ("Bさん", "メンバー", "Aさん"),
            ("Cさん", "メンバー", "Aさん"),
            ("Dさん", "リーダー", "Dさん"),
            ("Eさん", "メンバー", "Dさん"),
            ("Fさん", "メンバー", "Dさん"),
        ]
        for i, (name, role, leader) in enumerate(members, start=5):
            ws.cell(row=i, column=1, value=name)
            ws.cell(row=i, column=2, value=role)
            ws.cell(row=i, column=3, value=leader)
            
        # 名前定義の追加（メンバーリストの範囲）
        # 将来的にデータ入力規則（プルダウン）で使用することを想定
        member_range = f"Settings!$A$5:$A$100"
        defn = DefinedName("MEMBER_LIST", attr_text=member_range)
        self.wb.defined_names.add(defn)

    def _create_wbs_evm_sheet(self):
        """WBS_EVMシートを作成し、3フェーズ構造（担当者・リーダーを左端に配置）を構築する"""
        ws = self.wb.create_sheet("WBS_EVM")
        
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        header_font = Font(bold=True)
        alignment = Alignment(horizontal="center", vertical="center")
        
        # 1フェーズ12列構成。各フェーズの開始列を定義。
        # 4列目(D)から開始。No(1), ID(2), 名称(3) の後。
        phases = [
            ("作成", 4, 15),           # D(4)からO(15)
            ("レビュー実施", 16, 27),   # P(16)からAA(27)
            ("レビュー後修正", 28, 39)  # AB(28)からAM(39)
        ]
        
        for name, start_col, end_col in phases:
            ws.cell(row=1, column=start_col, value=name)
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            cell = ws.cell(row=1, column=start_col)
            cell.alignment = alignment
            cell.font = header_font
            cell.fill = header_fill

        # 2行目: カラムヘッダー
        base_cols = ["No", "機能ID", "機能名称"]
        for i, col_name in enumerate(base_cols, start=1):
            cell = ws.cell(row=2, column=i, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            
        # 担当・リーダーを左端に配置した列構成 (計12列)
        phase_cols = [
            "担当メンバー", "チームリーダー", # 0, 1
            "開始日予定", "終了日予定", "工数予定", # 2, 3, 4
            "開始日実績", "終了日実績", "工数実績", # 5, 6, 7
            "進捗率(%)", "PV (計画値)", "EV (出来高)", "AC (実績コスト)" # 8, 9, 10, 11
        ]
        
        for i, phase in enumerate(phases):
            start_col = phase[1]
            for j, col_name in enumerate(phase_cols):
                cell = ws.cell(row=2, column=start_col + j, value=col_name)
                cell.font = header_font
                cell.alignment = alignment
                cell.fill = header_fill

        # 3行目から100行目まで数式とスタイルを適用
        for row in range(3, 103):
            for i, phase in enumerate(phases):
                s_col = phase[1] # 各フェーズの開始列番号
                
                # 列記号の取得 (数式作成用)
                # s_col からの相対位置で指定
                plan_start = ws.cell(row=row, column=s_col+2).column_letter # 開始日予定
                plan_end = ws.cell(row=row, column=s_col+3).column_letter   # 終了日予定
                plan_cost = ws.cell(row=row, column=s_col+4).column_letter  # 工数予定
                actual_cost_input = ws.cell(row=row, column=s_col+7).column_letter # 工数実績
                progress = ws.cell(row=row, column=s_col+8).column_letter   # 進捗率(%)
                
                # 進捗率(%)セルの書式設定: カスタム書式「0"%"」により100入力で100%と表示
                progress_cell = ws.cell(row=row, column=s_col+8)
                progress_cell.number_format = '0"%"'
                progress_cell.alignment = alignment

                # PV (Planned Value / 計画値): s_col + 9
                # 今日の日付に基づき、予定期間の消化具合から算出
                pv_formula = f'=IF(TODAY()<{plan_start}{row}, 0, IF(TODAY()>{plan_end}{row}, {plan_cost}{row}, {plan_cost}{row} * (TODAY()-{plan_start}{row})/({plan_end}{row}-{plan_start}{row}+1)))'
                pv_cell = ws.cell(row=row, column=s_col+9, value=pv_formula)
                pv_cell.alignment = alignment
                
                # EV (Earned Value / 出来高): s_col + 10
                # 工数予定に進捗率を掛けて算出 (/100 はパーセント入力対応)
                ev_formula = f'={plan_cost}{row} * ({progress}{row}/100)'
                ev_cell = ws.cell(row=row, column=s_col+10, value=ev_formula)
                ev_cell.alignment = alignment
                
                # AC (Actual Cost / 実績コスト): s_col + 11
                # 入力された実績工数をそのまま参照
                ac_formula = f'={actual_cost_input}{row}'
                ac_cell = ws.cell(row=row, column=s_col+11, value=ac_formula)
                ac_cell.alignment = alignment

        # アラート（条件付き書式）の設定
        # PV列 (作成: M(13), レビュー: Y(25), 修正: AK(37)) に 0未満アラート(デモ用)
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        for col in ["M", "Y", "AK"]:
            ws.conditional_formatting.add(f"{col}3:{col}103", CellIsRule(operator='lessThan', formula=['0'], fill=red_fill))

        # 罫線の適用 (A1 から AM102 まで)
        self._apply_borders(ws, 1, 102, 1, 39)

    def _create_team_evm_sheet(self):
        """
        チームEVMシートを作成。
        リーダーごとに独立した集計テーブルを構築し、WBS_EVMシートの各フェーズ列を自動集計する。
        """
        ws = self.wb.create_sheet("チームEVM")
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # 集計対象のリーダーリスト
        leaders = ["Aさん", "Dさん"]
        
        current_row = 1
        for leader in leaders:
            start_row_block = current_row # 罫線囲み用の開始行
            
            # チームタイトル
            ws.cell(row=current_row, column=1, value=f"{leader}チーム").font = Font(bold=True, size=12)
            current_row += 1
            
            # --- 1. EVM指標集計テーブル ---
            ws.cell(row=current_row, column=1, value="EVM指標集計").font = header_font
            current_row += 1
            
            evm_headers = ["フェーズ", "PV (計画値)", "EV (出来高)", "AC (実績コスト)", "SV", "CV", "SPI", "CPI"]
            for i, h in enumerate(evm_headers, start=1):
                cell = ws.cell(row=current_row, column=i, value=h)
                cell.font = header_font
                cell.fill = header_fill
            current_row += 1
            
            # 各フェーズの指標列と、そのフェーズの責任者（リーダー）列の対応
            # 作成: D(4)-O(15) -> リーダー:E(5), PV:M(13), EV:N(14), AC:O(15)
            # レビュー: P(16)-AA(27) -> リーダー:Q(17), PV:Y(25), EV:Z(26), AC:AA(27)
            # 修正: AB(28)-AM(39) -> リーダー:AC(29), PV:AK(37), EV:AL(38), AC:AM(39)
            phases_info = [
                ("作成", "M", "N", "O", "E"), 
                ("レビュー実施", "Y", "Z", "AA", "Q"), 
                ("レビュー後修正", "AK", "AL", "AM", "AC")
            ]
            for phase_name, pv_col, ev_col, ac_col, leader_col in phases_info:
                ws.cell(row=current_row, column=1, value=phase_name)
                # SUMIFS(集計列, 条件範囲(リーダー列), 条件値)
                ws.cell(row=current_row, column=2, value=f'=SUMIFS(WBS_EVM!{pv_col}:{pv_col}, WBS_EVM!${leader_col}:${leader_col}, "{leader}")')
                ws.cell(row=current_row, column=3, value=f'=SUMIFS(WBS_EVM!{ev_col}:{ev_col}, WBS_EVM!${leader_col}:${leader_col}, "{leader}")')
                ws.cell(row=current_row, column=4, value=f'=SUMIFS(WBS_EVM!{ac_col}:{ac_col}, WBS_EVM!${leader_col}:${leader_col}, "{leader}")')
                # SV = EV - PV, CV = EV - AC
                ws.cell(row=current_row, column=5, value=f'=C{current_row}-B{current_row}')
                ws.cell(row=current_row, column=6, value=f'=C{current_row}-D{current_row}')
                # SPI = EV / PV, CPI = EV / AC (ゼロ除算回避)
                ws.cell(row=current_row, column=7, value=f'=IFERROR(C{current_row}/B{current_row}, 1)')
                ws.cell(row=current_row, column=8, value=f'=IFERROR(C{current_row}/D{current_row}, 1)')
                current_row += 1
            
            current_row += 1 # スペース
            
            # --- 2. タスクメトリクステーブル ---
            ws.cell(row=current_row, column=1, value="タスクメトリクス").font = header_font
            current_row += 1
            
            met_headers = ["フェーズ", "総数", "仕掛かり(予定)", "仕掛かり(実績)", "完了(予定)", "完了(実績)"]
            for i, h in enumerate(met_headers, start=1):
                cell = ws.cell(row=current_row, column=i, value=h)
                cell.font = header_font
                cell.fill = header_fill
            current_row += 1
            
            # メトリクス算出用の参照列 (責任者、予定工数、終了日実績)
            # 作成: 予定工数:H(8), 終了日実績:K(11), リーダー:E(5)
            # レビュー: 予定工数:T(20), 終了日実績:W(23), リーダー:Q(17)
            # 修正: 予定工数:AF(32), 終了日実績:AI(35), リーダー:AC(29)
            metrics_info = [
                ("作成", "H", "K", "E"), 
                ("レビュー実施", "T", "W", "Q"), 
                ("レビュー後修正", "AF", "AI", "AC")
            ]
            for p_name, cost_col, end_act_col, lead_col in metrics_info:
                ws.cell(row=current_row, column=1, value=p_name)
                # 総数: リーダーが一致し、予定工数が入っている行をカウント
                ws.cell(row=current_row, column=2, value=f'=COUNTIFS(WBS_EVM!${lead_col}:${lead_col}, "{leader}", WBS_EVM!${cost_col}:${cost_col}, ">0")')
                # 完了(実績): リーダーが一致し、終了日実績が入力されている行をカウント
                ws.cell(row=current_row, column=6, value=f'=COUNTIFS(WBS_EVM!${lead_col}:${lead_col}, "{leader}", WBS_EVM!${end_act_col}:${end_act_col}, "<>")')
                current_row += 1

            # チームブロック全体に罫線を適用
            self._apply_borders(ws, start_row_block, current_row - 1, 1, 8)
            current_row += 2 # 次のチームへの間隔

    def _create_individual_sv_sheet(self):
        """個人SVシートを作成し、メンバーごとの日次追跡テーブルを構築する"""
        ws = self.wb.create_sheet("個人SV")
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center")
        
        # 1行目: カラムヘッダー
        headers = ["No", "メンバー", "前日", "当日", "日付"]
        for i, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            
        # 日付エリアのサンプル (横に広がる)
        dates = ["5月3日(日)", "5月4日(月)", "5月5日(火)"]
        for i, d in enumerate(dates, start=5): # E列(5)から開始
            cell = ws.cell(row=2, column=i, value=d)
            cell.font = header_font
            cell.alignment = alignment

        # メンバーリストの転記 (Settingsシートを参照する想定だが、ここではサンプルを直接配置)
        members = ["Aさん", "Bさん", "Cさん", "Dさん", "Eさん", "Fさん"]
        for i, name in enumerate(members, start=3): # 3行目から開始
            ws.cell(row=i, column=1, value=i-2) # No
            ws.cell(row=i, column=2, value=name) # メンバー名
            
        # 罫線の適用
        self._apply_borders(ws, 1, len(members) + 2, 1, 7) # サンプルとして7列目まで

    def generate(self, config=None):
        """全てのシートを生成し、Excelファイルを保存する"""
        self.config = config
        if config:
            self._create_settings_sheet()
            self._create_dynamic_wbs_sheet(config)
            # チームEVM等は構造が複雑なため、現在はデフォルト構造のみ対応
            # self._create_team_evm_sheet() 
        else:
            self._create_settings_sheet()
            self._create_wbs_evm_sheet()
            self._create_team_evm_sheet()
            self._create_individual_sv_sheet()
        
        os.makedirs(os.path.dirname(self.output_path), exist_ok=True)
        self.wb.save(self.output_path)
        print(f"Template generated at: {self.output_path}")

    def _create_dynamic_wbs_sheet(self, config):
        """JSON設定に基づき、動的なフェーズ・カラム構成でWBSシートを構築する"""
        sheet_name = config.get("sheet_name", "WBS_EVM")
        ws = self.wb.create_sheet(sheet_name)
        
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        header_font = Font(bold=True)
        alignment = Alignment(horizontal="center", vertical="center")
        
        # 1. フェーズ見出しの配置
        for phase in config["columns"]["phases"]:
            # そのフェーズに含まれる列の最小/最大インデックスを特定
            indices = [info["index"] for info in phase["mapping"].values()]
            if not indices: continue
            
            start_col = min(indices) + 1
            end_col = max(indices) + 1
            
            ws.cell(row=1, column=start_col, value=phase["name"])
            if start_col < end_col:
                ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            
            cell = ws.cell(row=1, column=start_col)
            cell.alignment = alignment
            cell.font = header_font
            cell.fill = header_fill

        # 2. カラムヘッダー（2行目）の配置
        # columns["all"] に基づいて全ての列を配置
        max_idx = 0
        for col in config["columns"]["all"]:
            idx = col["index"] + 1
            max_idx = max(max_idx, idx)
            cell = ws.cell(row=2, column=idx, value=col["name"])
            cell.font = header_font
            cell.alignment = alignment
            cell.fill = header_fill

        # 3. スタイルの適用（罫線）
        self._apply_borders(ws, 1, 102, 1, max_idx, outer_medium=True)

        # 4. データ行への数式適用 (3行目から102行目)
        for row in range(3, 103):
            for phase in config["columns"]["phases"]:
                mapping = phase["mapping"]
                
                # 役割ごとの列記号を取得
                def get_col(role):
                    if role in mapping:
                        return get_column_letter(mapping[role]["index"] + 1)
                    return None

                p_start = get_col("plan_start")
                p_end = get_col("plan_end")
                p_effort = get_col("plan_effort")
                a_effort = get_col("actual_effort")
                progress = get_col("progress")
                pv = get_col("pv")
                ev = get_col("ev")
                ac = get_col("ac")

                # 進捗率セルの書式設定
                if progress:
                    prog_cell = ws.cell(row=row, column=mapping["progress"]["index"] + 1)
                    prog_cell.number_format = '0"%"'
                    prog_cell.alignment = alignment

                # PV 数式
                if pv and p_start and p_end and p_effort:
                    formula = f'=IF(TODAY()<{p_start}{row}, 0, IF(TODAY()>{p_end}{row}, {p_effort}{row}, {p_effort}{row} * (TODAY()-{p_start}{row})/({p_end}{row}-{p_start}{row}+1)))'
                    ws.cell(row=row, column=mapping["pv"]["index"] + 1, value=formula).alignment = alignment

                # EV 数式
                if ev and p_effort and progress:
                    formula = f'={p_effort}{row}*({progress}{row}/100)'
                    ws.cell(row=row, column=mapping["ev"]["index"] + 1, value=formula).alignment = alignment

                # AC 数式
                if ac and a_effort:
                    formula = f'={a_effort}{row}'
                    ws.cell(row=row, column=mapping["ac"]["index"] + 1, value=formula).alignment = alignment

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="WBS/EVM マスターテンプレート生成スクリプト")
    parser.add_argument("--config", help="WBS構造を定義したJSONファイルのパス")
    parser.add_argument("--output", default="templates/master_template.xlsx", help="生成するエクセルの出力パス")
    parser.add_argument("--sheet-name", help="WBSシート名の指定（config内の設定を上書き）")
    
    args = parser.parse_args()
    
    config = None
    if args.config:
        if not os.path.exists(args.config):
            print(f"Error: Config file not found: {args.config}")
            exit(1)
        with open(args.config, "r", encoding="utf-8") as f:
            config = json.load(f)
        
        if args.sheet_name:
            config["sheet_name"] = args.sheet_name

    generator = TemplateGenerator(args.output)
    generator.generate(config=config)
