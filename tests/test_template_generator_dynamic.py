import pytest
import os
import openpyxl
from scripts.generate_master_template import TemplateGenerator

def test_generate_from_custom_config(tmp_path):
    """
    JSON 設定に従って、任意のフェーズとカラムを持つエクセルが生成されることを検証。
    """
    output_path = str(tmp_path / "custom_master.xlsx")
    config = {
        "sheet_name": "Custom_WBS",
        "header_row": 2,
        "columns": {
            "common": {
                "id": {"index": 0, "name": "No."},
                "name": {"index": 1, "name": "項目名"}
            },
            "phases": [
                {
                    "name": "独自フェーズA",
                    "mapping": {
                        "plan_start": {"index": 2, "name": "開始"},
                        "plan_effort": {"index": 3, "name": "工数"}
                    }
                }
            ],
            "all": [
                {"name": "No.", "index": 0, "role": "id"},
                {"name": "項目名", "index": 1, "role": "name"},
                {"name": "開始", "index": 2, "role": "plan_start"},
                {"name": "工数", "index": 3, "role": "plan_effort"}
            ]
        }
    }
    
    gen = TemplateGenerator(output_path)
    # 現在の TemplateGenerator は config 引数を持っていないため、ここで失敗するはず
    gen.generate(config=config)
    
    assert os.path.exists(output_path)
    wb = openpyxl.load_workbook(output_path)
    assert "Custom_WBS" in wb.sheetnames
    
    ws = wb["Custom_WBS"]
    assert ws.cell(row=1, column=3).value == "独自フェーズA"
    assert ws.cell(row=2, column=1).value == "No."
    assert ws.cell(row=2, column=2).value == "項目名"
    assert ws.cell(row=2, column=3).value == "開始"

def test_dynamic_formula_references(tmp_path):
    """
    動的に配置された列に対して、正しい相対参照の数式が生成されていることを検証。
    """
    output_path = str(tmp_path / "formula_test.xlsx")
    # 工数予定(index:2), 進捗(index:3) の位置に配置
    config = {
        "sheet_name": "WBS_EVM",
        "header_row": 2,
        "columns": {
            "common": {"id": {"index": 0, "name": "ID"}, "name": {"index": 1, "name": "Name"}},
            "phases": [
                {
                    "name": "Phase1",
                    "mapping": {
                        "plan_effort": {"index": 2, "name": "予定工数"},
                        "progress": {"index": 3, "name": "進捗率"},
                        "ev": {"index": 4, "name": "出来高(EV)"}
                    }
                }
            ],
            "all": [
                {"name": "ID", "index": 0, "role": "id"},
                {"name": "Name", "index": 1, "role": "name"},
                {"name": "予定工数", "index": 2, "role": "plan_effort"},
                {"name": "進捗率", "index": 3, "role": "progress"},
                {"name": "出来高(EV)", "index": 4, "role": "ev"}
            ]
        }
    }
    
    gen = TemplateGenerator(output_path)
    gen.generate(config=config)
    
    wb = openpyxl.load_workbook(output_path)
    ws = wb["WBS_EVM"]
    
    # 3行目の EV 数式を検証
    # 予定工数は C列 (3), 進捗率は D列 (4), EVは E列 (5)
    # 期待される数式: =C3*(D3/100)
    ev_formula = ws.cell(row=3, column=5).value
    assert ev_formula == "=C3*(D3/100)"

def test_robustness_pv_formula_and_header_row(tmp_path):
    """
    1. PV数式がゼロ除算（開始日=終了日）を回避しているか検証
    2. ヘッダー行位置(header_row)がJSONに従って動的に変わるか検証
    """
    output_path = str(tmp_path / "robustness_test.xlsx")
    config = {
        "sheet_name": "WBS",
        "header_row": 5, # 5行目をヘッダーにする
        "columns": {
            "common": {"id": {"index": 0, "name": "ID"}},
            "phases": [
                {
                    "name": "P1",
                    "mapping": {
                        "plan_start": {"index": 1, "name": "S"},
                        "plan_end": {"index": 2, "name": "E"},
                        "plan_effort": {"index": 3, "name": "C"},
                        "pv": {"index": 4, "name": "PV"}
                    }
                }
            ],
            "all": [
                {"name": "ID", "index": 0, "role": "id"},
                {"name": "S", "index": 1, "role": "plan_start"},
                {"name": "E", "index": 2, "role": "plan_end"},
                {"name": "C", "index": 3, "role": "plan_effort"},
                {"name": "PV", "index": 4, "role": "pv"}
            ]
        }
    }

    gen = TemplateGenerator(output_path)
    gen.generate(config=config)

    wb = openpyxl.load_workbook(output_path)
    ws = wb["WBS"]

    # 検証1: ヘッダー位置 (header_row=5 なので、カラム名は 5行目、フェーズ名は 4行目にあるべき)
    assert ws.cell(row=4, column=2).value == "P1"
    assert ws.cell(row=5, column=1).value == "ID"

    # 検証2: PV数式のゼロ除算保護
    # 期待される修正後の数式（例）: 分母に +1 を含むか、MAX(1, ...) で保護されている
    # 既存は (E-S+1) だったので、S=E なら分母は 1 になるはず。
    # レビュアーの指摘は「逆転時 (E<S)」に 0 や 負 になるリスク。
    # 安全な数式: ... / MAX(1, E-S+1)
    pv_formula = ws.cell(row=6, column=5).value
    assert "MAX(1," in pv_formula

def test_dynamic_conditional_formatting(tmp_path):
    """
    動的生成時に、PV列に対して条件付き書式が設定されていることを検証。
    """
    output_path = str(tmp_path / "cf_test.xlsx")
    config = {
        "sheet_name": "WBS",
        "columns": {
            "common": {"id": {"index": 0, "name": "ID"}},
            "phases": [{"name": "P1", "mapping": {"pv": {"index": 1, "name": "PV"}}}],
            "all": [{"name": "ID", "index": 0, "role": "id"}, {"name": "PV", "index": 1, "role": "pv"}]
        }
    }
    gen = TemplateGenerator(output_path)
    gen.generate(config=config)
    
    wb = openpyxl.load_workbook(output_path)
    ws = wb["WBS"]
    
    # 条件付き書式が存在することを確認 (B列3行目から102行目まで)
    assert len(ws.conditional_formatting) > 0


def test_cli_integration(tmp_path):
    """
    コマンドライン引数 --config を使用してエクセルが生成されることを検証。
    """
    import json
    import subprocess
    import sys
    
    config = {
        "sheet_name": "CLI_Sheet",
        "header_row": 2,
        "columns": {
            "common": {"id": {"index": 0, "name": "ID"}, "name": {"index": 1, "name": "Name"}},
            "phases": [{"name": "P1", "mapping": {"progress": {"index": 2, "name": "P"}}}],
            "all": [{"name": "ID", "index": 0, "role": "id"}, {"name": "Name", "index": 1, "role": "name"}, {"name": "P", "index": 2, "role": "progress"}]
        }
    }
    config_path = tmp_path / "cli_config.json"
    with open(config_path, "w") as f:
        json.dump(config, f)
        
    output_path = tmp_path / "cli_out.xlsx"
    
    # スクリプトの実行 (python scripts/generate_master_template.py --config ... --output ...)
    cmd = [
        sys.executable, 
        "scripts/generate_master_template.py", 
        "--config", str(config_path), 
        "--output", str(output_path)
    ]
    
    result = subprocess.run(cmd, capture_output=True, text=True)
    assert result.returncode == 0
    assert os.path.exists(output_path)
    
    wb = openpyxl.load_workbook(output_path)
    assert "CLI_Sheet" in wb.sheetnames
