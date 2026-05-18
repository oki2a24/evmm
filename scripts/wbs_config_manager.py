from typing import Dict, List, Optional, Any
import pandas as pd
import openpyxl
import re
import os
import json

class WBSConfigManager:
    """
    WBSエクセルの構造を解析し、役割ベースのカラムマッピングを管理するクラス。

    【設計背景 (Architecture Rationale)】
    本プロジェクトでは「Excel-as-Master」の思想に基づき、人間が自由に編集したエクセルを
    プログラムが理解する必要があります。このクラスは「翻訳者」として、エクセルの
    物理的な列名・位置を、プログラムが扱う抽象的な「役割（Role）」へと変換します。
    これにより、プロジェクトごとにテンプレートが微修正されても、ロジックを壊さずに
    運用を継続することを可能にしています。
    """
    
    DEFAULT_ROLES = {
        "id": [r"機能ID", r"ID", r"識別子"],
        "name": [r"機能名称", r"タスク名", r"名称", r"項目名"],
        "plan_start": [r"開始日予定", r"予定開始日", r"開始予定", r"作業開始", r"着手日"],
        "plan_end": [r"終了日予定", r"予定終了日", r"終了予定", r"期限", r"デッドライン", r"完了予定"],
        "plan_effort": [r"工数予定", r"予定工数", r"見積工数", r"予定人日"],
        "actual_start": [r"開始日実績", r"実績開始日", r"開始実績", r"着手実績"],
        "actual_end": [r"終了日実績", r"実績終了日", r"終了実績", r"完了実績"],
        "actual_effort": [r"工数実績", r"実績工数", r"実工数"],
        "progress": [r"進捗率", r"進捗", r"達成率"],
        "pv": [r"PV", r"計画値"],
        "ev": [r"EV", r"出来高"],
        "ac": [r"AC", r"実績コスト"]
    }


    def __init__(self, file_path: str, sheet_name: str = 'WBS_EVM'):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.config_path = os.path.join(os.path.dirname(os.path.abspath(file_path)), "wbs_structure.json")
        self.config: Optional[Dict[str, Any]] = None

    def load_or_infer(self, interactive: bool = True) -> Dict[str, Any]:
        """
        設定ファイルを読み込む。存在しないか構造が変更されている場合は推論と対話を行う。
        """
        if os.path.exists(self.config_path):
            with open(self.config_path, "r", encoding="utf-8") as f:
                self.config = json.load(f)
            
            # 構造の生存確認
            if self._validate_current_config():
                return self.config
            else:
                print("WBS構造が変更されたようです。再マッピングを行います。")

        # 推論実行
        inferred = self.infer_structure()
        
        if interactive:
            self.config = self._interactive_refine(inferred)
            self.save_config()
        else:
            self.config = inferred
            
        return self.config

    def _validate_current_config(self) -> bool:
        """
        現在の設定が実際のエクセルファイルと一致するか確認する。
        """
        if not self.config:
            return False
            
        wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
        if self.sheet_name not in wb.sheetnames:
            return False
        
        ws = wb[self.sheet_name]
        h_row = self.config["header_row"]
        
        # 1. 共通項目のチェック
        common = self.config["columns"]["common"]
        for role, info in common.items():
            cell_val = str(ws.cell(row=h_row, column=info["index"] + 1).value)
            if info["name"] not in cell_val:
                return False
        
        # 2. フェーズ項目のチェック
        for p in self.config["columns"]["phases"]:
            for role, info in p["mapping"].items():
                cell_val = str(ws.cell(row=h_row, column=info["index"] + 1).value)
                if info["name"] not in cell_val:
                    return False
                
        return True

    def infer_structure(self) -> Dict[str, Any]:
        """
        エクセルのヘッダーをスキャンし、構造を推論する。

        【推論の規律】
        1. ID列の特定: 最初に「機能ID」等のキーワードを含む列を探し、そこをヘッダー行とみなします。
        2. フェーズの分離: ヘッダー行の上方向にスキャンして、結合セルや単一セルで書かれた「作成」等の
           フェーズ名を取得し、以降のカラムをそのフェーズに所属させます。
        3. 役割の割当: 各列名に対して正規表現マッチングを行い、最も近い役割を割り当てます。
        """
        # openpyxl を使用して物理ファイルをスキャン。
        # ※テスト時にも物理的な Zip (xlsx) 構造を要求するため、テストコード側で
        #   物理ファイルの生成が必要になる点に注意してください。
        wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
        if self.sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{self.sheet_name}' not found.")
        
        ws = wb[self.sheet_name]
        # レビュアー指摘: スキャン範囲を拡張 (20行)
        rows = list(ws.iter_rows(min_row=1, max_row=20, values_only=True))
        
        header_row_idx = -1
        for i, row in enumerate(rows):
            if any(re.search(p, str(cell)) for cell in row if cell for p in self.DEFAULT_ROLES["id"]):
                header_row_idx = i
                break
        
        if header_row_idx == -1:
            raise ValueError("Could not find header row (ID column missing).")

        header_row = rows[header_row_idx]
        
        # 2. フェーズ行の特定 (上方向にスキャンして非空行を探す)
        phase_row_idx = -1
        for i in range(header_row_idx - 1, -1, -1):
            if any(str(cell).strip() for cell in rows[i] if cell):
                phase_row_idx = i
                break
        
        phase_row = rows[phase_row_idx] if phase_row_idx != -1 else [None] * len(header_row)

        config = {
            "sheet_name": self.sheet_name,
            "header_row": header_row_idx + 1,
            "data_start_row": header_row_idx + 2,
            "columns": {
                "common": {},
                "phases": [],
                "all": []  # 全ての列情報を順序通りに保持
            }
        }

        current_phase = None
        phase_data = None

        for i, (col_name, p_name) in enumerate(zip(header_row, phase_row)):
            if col_name is None: 
                # 空列も位置保持のために記録
                config["columns"]["all"].append({"name": None, "index": i})
                continue
                
            col_name = str(col_name)
            config["columns"]["all"].append({"name": col_name, "index": i})

            if p_name and str(p_name).strip():
                if phase_data:
                    config["columns"]["phases"].append(phase_data)
                current_phase = str(p_name).strip()
                phase_data = {"name": current_phase, "mapping": {}}

            matched_role = None
            for role, patterns in self.DEFAULT_ROLES.items():
                if any(re.search(p, col_name) for p in patterns):
                    matched_role = role
                    break
            
            if matched_role:
                info = {"name": col_name, "index": i}
                if matched_role in ["id", "name"]:
                    config["columns"]["common"][matched_role] = info
                elif phase_data:
                    if matched_role not in phase_data["mapping"]:
                        phase_data["mapping"][matched_role] = info

        if phase_data:
            config["columns"]["phases"].append(phase_data)

        return config

    def _interactive_refine(self, config: Dict[str, Any]) -> Dict[str, Any]:
        """
        推論結果をユーザーに提示し、必要に応じて修正を受け付ける。
        """
        print("\n--- WBS構造の解析結果 ---")
        print(f"シート名: {config['sheet_name']}")
        print(f"ヘッダー行: {config['header_row']}")
        
        print("\n[共通項目]")
        for role, info in config["columns"]["common"].items():
            print(f"  - {role}: {info['name']} (列 {info['index']+1})")
            
        print("\n[フェーズ構成]")
        for p in config["columns"]["phases"]:
            print(f"  フェーズ: {p['name']}")
            for role, info in p["mapping"].items():
                print(f"    - {role}: {info['name']} (列 {info['index']+1})")
        
        ans = input("\nこのマッピングでよろしいですか？ (Y/n) ※nを入力すると手動設定ガイドを表示します: ").lower()
        if ans == 'n':
            print("\n【手動設定ガイド】")
            print("推論が外れている場合は、エクセルのカラム名を以下のいずれかに含めるように修正してください。")
            for role, patterns in self.DEFAULT_ROLES.items():
                print(f"  - {role:12}: {', '.join(patterns)}")
            print("\n名称を修正後、再度実行してください。")
            # 本来はここで詳細な修正UIを提供すべきだが、まずは名称変更を促すことでSSOTを維持
            
        return config


    def save_config(self):
        """設定をJSONに保存する。"""
        if not self.config: return
        with open(self.config_path, "w", encoding="utf-8") as f:
            json.dump(self.config, f, indent=2, ensure_ascii=False)
        print(f"設定を保存しました: {self.config_path}")

    def get_column_index(self, role: str, phase_idx: int = 0) -> int:
        """
        役割とフェーズインデックスから列インデックス(0-based)を取得する。

        【ガードレール】
        インデックスを直接返さずこのメソッドを経由させることで、カラムの欠落や
        フェーズ名の変更に対して、実行時に具体的なエラーメッセージを提示できます。
        """
        if not self.config:
            self.load_or_infer(interactive=False)
            
        if role in self.config["columns"]["common"]:
            return self.config["columns"]["common"][role]["index"]
        
        if phase_idx < len(self.config["columns"]["phases"]):
            mapping = self.config["columns"]["phases"][phase_idx]["mapping"]
            if role in mapping:
                return mapping[role]["index"]
            
            phase_name = self.config["columns"]["phases"][phase_idx]["name"]
            raise ValueError(f"必須項目 '{role}' がフェーズ '{phase_name}' (インデックス {phase_idx}) で見つかりません。エクセルのカラム名を確認してください。")
        
        raise ValueError(f"共通項目またはフェーズ {phase_idx} において、役割 '{role}' が定義されていません。")
