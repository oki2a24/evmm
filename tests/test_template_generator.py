import os
import pytest
from openpyxl import load_workbook

from scripts.generate_master_template import TemplateGenerator

# 生成されるテンプレートのパス
TEMPLATE_PATH = "templates/master_template.xlsx"

def test_template_file_exists():
    """
    ステップ 1: テンプレートファイルが物理的に生成されることを検証する。
    """
    if os.path.exists(TEMPLATE_PATH):
        os.remove(TEMPLATE_PATH)

    # 最小限の実装を呼び出す
    generator = TemplateGenerator(TEMPLATE_PATH)
    generator.generate()

    assert os.path.exists(TEMPLATE_PATH), "テンプレートファイルが生成されていません。"

def test_settings_sheet_structure():
    """
    タスク 2: Settings シートの構造を検証する。
    """
    generator = TemplateGenerator(TEMPLATE_PATH)
    generator.generate()
    
    wb = load_workbook(TEMPLATE_PATH)
    assert "Settings" in wb.sheetnames, "Settings シートが存在しません。"
    
    ws = wb["Settings"]
    # 期待される見出しの検証
    assert ws["A1"].value == "プロジェクト基本情報", "A1に見出しがありません。"
    assert ws["A3"].value == "メンバーリスト", "A3に見出しがありません。"

def test_wbs_evm_sheet_structure():
    """
    タスク 3: WBS_EVM シートの 3フェーズ構造と日本語併記を検証する。
    """
    generator = TemplateGenerator(TEMPLATE_PATH)
    generator.generate()
    
    wb = load_workbook(TEMPLATE_PATH)
    assert "WBS_EVM" in wb.sheetnames, "WBS_EVM シートが存在しません。"
    
    ws = wb["WBS_EVM"]
    
    # 3フェーズヘッダーの確認 (結合されていることを想定)
    # 作成: D1(4), レビュー実施: P1(16), レビュー後修正: AB1(28)
    assert ws["D1"].value == "作成", "作成フェーズのヘッダーがありません。"
    assert ws["P1"].value == "レビュー実施", "レビュー実施フェーズのヘッダーがありません。"
    assert ws["AB1"].value == "レビュー後修正", "レビュー後修正フェーズのヘッダーがありません。"

    # 日本語併記の確認
    # 各フェーズの列名を確認 (1フェーズ12列構成)
    expected_cols = [
        "開始日予定", "終了日予定", "工数予定", "開始日実績", "終了日実績", "工数実績",
        "進捗率(%)", "PV (計画値)", "EV (出来高)", "AC (実績コスト)", "担当メンバー", "チームリーダー"
    ]
    all_col_values = [ws.cell(row=2, column=c).value for c in range(1, 45)]
    for col_name in expected_cols:
        assert col_name in all_col_values, f"{col_name} の列が見当たりません。"

def test_wbs_evm_formulas_and_alerts():
    """
    タスク 4: WBS_EVM シートの計算式とアラート設定を検証する。
    """
    generator = TemplateGenerator(TEMPLATE_PATH)
    generator.generate()

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["WBS_EVM"]

    # 3行目の数式を検証 (作成フェーズ: D列(4)〜O列(15) の12列構成)
    # PV (計画値): K列(11), EV (出来高): L列(12), AC (実績コスト): M列(13)
    assert ws.cell(row=3, column=11).data_type == 'f', "11列目(PV) に数式が設定されていません。"
    assert ws.cell(row=3, column=12).data_type == 'f', "12列目(EV) に数式が設定されていません。"
    assert ws.cell(row=3, column=13).data_type == 'f', "13列目(AC) に数式が設定されていません。"
    
    # 条件付き書式の存在確認 (SPI/CPI 等)
    assert len(ws.conditional_formatting) > 0, "条件付き書式が設定されていません。"

def test_team_evm_sheet_summary():
    """
    タスク 5: チームEVM シートの集計テーブルと数式を検証する。
    """
    generator = TemplateGenerator(TEMPLATE_PATH)
    generator.generate()
    
    wb = load_workbook(TEMPLATE_PATH)
    assert "チームEVM" in wb.sheetnames, "チームEVM シートが存在しません。"
    
    ws = wb["チームEVM"]
    
    # チームAの集計ブロックを確認 (設計に基づきAさんチームが最初に来る)
    assert "Aさんチーム" in [ws.cell(row=r, column=1).value for r in range(1, 20)], "Aさんチームの集計ブロックがありません。"
    
    # SUMIFS / COUNTIFS 数式の存在確認 (メトリクステーブルのどこか)
    formulas = [ws.cell(row=r, column=c).value for r in range(1, 40) for c in range(1, 20) if ws.cell(row=r, column=c).data_type == 'f']
    assert any("SUMIFS" in str(f) or "COUNTIFS" in str(f) for f in formulas), "集計用の SUMIFS/COUNTIFS 数式が見当たりません。"
